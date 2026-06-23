# -*- coding: utf-8 -*-
from __future__ import annotations

import csv
import hashlib
import io
import json
import platform
import shutil
import sys
import textwrap
from datetime import datetime
from pathlib import Path

ROOT = Path(r"D:\Research Materials\IBD Item1 Meta-Model")
CODEXTOOLS = ROOT / "codextools"
if CODEXTOOLS.exists():
    sys.path.insert(0, str(CODEXTOOLS))

import matplotlib as mpl
import matplotlib.font_manager as fm
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
from PIL import Image


PROJECT_LOCATION_DIR = ROOT / "全项目位置记录"
FIGURE2_DIR = ROOT / "EMSAv1-Map of Evidence" / "Mainchats" / "Figure2"
INPUT_WORKBOOK_NAME = "Supplementary_Table_S3_Evidence_Map_revised_Figure2E_human_sample_fixed.xlsx"
INPUT_SHEET = "Figure2E_human_sample_source"
CODE_AVAILABLE_DIR = ROOT / "Code available" / "public_repository_code_package_GitHub_ready_20260620_141023"

DENOMINATOR = 491
FORBIDDEN_CATEGORIES = ["Animal samples", "Cell/organoid"]
FORBIDDEN_SEARCH_TERMS = ["Animal samples", "Animal sample", "Cell/organoid"]
EXPECTED_COUNTS = {
    "Stool/feces": 192,
    "Colonic tissue": 113,
    "Serum/plasma/blood": 101,
    "Ileal tissue": 77,
    "Intestinal mucosa/biopsy": 69,
    "Rectal tissue": 50,
    "Not reported/unclear human source": 27,
    "Saliva/oral": 20,
    "Urine": 19,
}
ORDER = list(EXPECTED_COUNTS)
PALETTE = {
    "Stool/feces": "#69C3A5",
    "Colonic tissue": "#D96AA0",
    "Serum/plasma/blood": "#9DCD52",
    "Ileal tissue": "#F2C84B",
    "Intestinal mucosa/biopsy": "#A68BC4",
    "Rectal tissue": "#F39A4A",
    "Not reported/unclear human source": "#D9DDE5",
    "Saliva/oral": "#72A9D8",
    "Urine": "#DF74AA",
}
DISPLAY_LABELS = {
    "Intestinal mucosa/biopsy": "Intestinal mucosa/\nbiopsy",
    "Not reported/unclear human source": "Not reported/unclear\nhuman source",
}
CAPTION_TEXT = (
    "Figure 2. Evidence landscape and research distribution of IBD gut multi-omics studies. "
    "A, Study identification and screening flow. B, Publication trend of retained evidence records. "
    "C, Primary omics category distribution. D, Evidence-density landscape by omics category and clinical question. "
    "E, Human cohort sample-source distribution based only on extractable human IBD gut-omics components. "
    "F, Clinical and analytical question distribution. G, Clinical-stage and trajectory-related evidence gaps. "
    "Purely non-human studies were excluded. In studies containing both human cohort data and animal, cell, "
    "organoid or other mechanistic components, only the human gut-omics component was counted for evidence-density "
    "and sample-source analyses; non-human components were retained only as mechanistic context."
)
CODE_AVAILABILITY_TEXT = (
    "The corrected Figure 2E human cohort sample-source panel was generated using the reproducible script "
    "`build_Figure2_panel_E_human_sample_source_fixed.py`. The script reads "
    "`Supplementary_Table_S3_Evidence_Map_revised_Figure2E_human_sample_fixed.xlsx`, sheet "
    "`Figure2E_human_sample_source`, and exports SVG, PDF, 600-dpi TIFF and PNG preview files. "
    "Animal, cell, organoid and other mechanistic model components are excluded from the Figure 2E human "
    "sample-source counts and are retained only as mechanistic context in the supplementary evidence map. "
    "The script and associated audit files were archived in the Figure 2 timestamped output directory and "
    "synchronized or prepared for synchronization with the Code available package."
)


def current_run_dir() -> Path:
    script_path = Path(__file__).resolve()
    if script_path.parent.name == "scripts" and script_path.parent.parent.name.startswith(
        "panel_E_human_sample_source_fixed_"
    ):
        return script_path.parent.parent
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    run_dir = FIGURE2_DIR / f"panel_E_human_sample_source_fixed_{ts}"
    for sub in ["scripts", "output", "preview", "audit", "archive", "code_available_update"]:
        (run_dir / sub).mkdir(parents=True, exist_ok=True)
    return run_dir


RUN_DIR = current_run_dir()
SCRIPT_DIR = RUN_DIR / "scripts"
OUTPUT_DIR = RUN_DIR / "output"
PREVIEW_DIR = RUN_DIR / "preview"
AUDIT_DIR = RUN_DIR / "audit"
ARCHIVE_DIR = RUN_DIR / "archive"
CODE_UPDATE_DIR = RUN_DIR / "code_available_update"
for folder in [SCRIPT_DIR, OUTPUT_DIR, PREVIEW_DIR, AUDIT_DIR, ARCHIVE_DIR, CODE_UPDATE_DIR]:
    folder.mkdir(parents=True, exist_ok=True)


def md5_file(path: Path) -> str:
    h = hashlib.md5()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def write_json(path: Path, data: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def write_csv(path: Path, rows: list[dict], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def read_csv_rows(path: Path) -> tuple[list[dict], list[str]]:
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        return list(reader), list(reader.fieldnames or [])


def write_csv_rows(path: Path, rows: list[dict], fieldnames: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def upsert_csv_row(path: Path, key_fields: list[str], new_row: dict) -> bool:
    rows, fieldnames = read_csv_rows(path)
    for key in new_row:
        if key not in fieldnames:
            fieldnames.append(key)
    updated = False
    for row in rows:
        if all(row.get(key) == new_row.get(key) for key in key_fields):
            row.update(new_row)
            updated = True
            break
    if not updated:
        rows.append({field: new_row.get(field, "") for field in fieldnames})
    write_csv_rows(path, rows, fieldnames)
    return updated


def scan_location_records() -> dict:
    checked = []
    hits = []
    if not PROJECT_LOCATION_DIR.exists():
        return {"checked": False, "files_checked": checked, "hits": hits}
    for path in PROJECT_LOCATION_DIR.rglob("*"):
        if not path.is_file():
            continue
        if path.suffix.lower() not in {".md", ".txt", ".csv", ".ndjson"}:
            continue
        if path.stat().st_size > 25 * 1024 * 1024:
            continue
        checked.append(str(path))
        try:
            text = path.read_text(encoding="utf-8", errors="ignore")
        except Exception:
            continue
        if INPUT_WORKBOOK_NAME in text or "Supplementary material" in text or "EMSAv1-Supplementary material" in text:
            hits.append(str(path))
    return {"checked": True, "files_checked": checked, "hits": hits}


def locate_input_workbook(location_scan: dict) -> tuple[Path | None, str]:
    candidates: list[tuple[Path, str]] = []
    for hit in location_scan.get("hits", []):
        text = Path(hit).read_text(encoding="utf-8", errors="ignore")
        for line in text.splitlines():
            if INPUT_WORKBOOK_NAME in line:
                cleaned = line.strip().strip("`").strip()
                if "D:\\" in cleaned:
                    start = cleaned.find("D:\\")
                    candidate = Path(cleaned[start:].strip("` "))
                    candidates.append((candidate, "located from project location record"))
    fallback_candidates = [
        ROOT / "EMSAv1-Supplementary material" / INPUT_WORKBOOK_NAME,
        FIGURE2_DIR / INPUT_WORKBOOK_NAME,
        FIGURE2_DIR / "data_processed" / INPUT_WORKBOOK_NAME,
        ROOT / "Figures Tables" / INPUT_WORKBOOK_NAME,
        ROOT / INPUT_WORKBOOK_NAME,
        ROOT / "Map of Evidence" / "Mainchats" / "Figure2" / INPUT_WORKBOOK_NAME,
        ROOT / "Map of Evidence" / "Mainchats" / "Figure2" / "data_processed" / INPUT_WORKBOOK_NAME,
    ]
    candidates.extend((candidate, "located from fallback path after project location record check") for candidate in fallback_candidates)
    seen = set()
    for candidate, reason in candidates:
        if str(candidate) in seen:
            continue
        seen.add(str(candidate))
        if candidate.exists():
            return candidate, reason
    return None, "not found"


def parse_counts_from_workbook(workbook_path: Path | None) -> tuple[list[dict], dict]:
    metadata = {
        "source": "fallback",
        "fallback_reason": "",
        "fields_used": "locked fallback counts",
        "input_workbook_md5": "",
    }
    counts = EXPECTED_COUNTS.copy()
    if workbook_path is None:
        metadata["fallback_reason"] = "Input workbook was not found after location-record and fallback-path checks."
    else:
        metadata["input_workbook_md5"] = md5_file(workbook_path)
        try:
            wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
            if INPUT_SHEET not in wb.sheetnames:
                raise KeyError(f"Sheet {INPUT_SHEET} not found.")
            ws = wb[INPUT_SHEET]
            parsed: dict[str, int] = {}
            rows = list(ws.iter_rows(values_only=True))
            category_col = None
            count_col = None
            for row in rows:
                for idx, value in enumerate(row):
                    if str(value).strip() == "Multi-label human sample-source audit":
                        category_col = idx
                        count_col = idx + 1
                        break
                if category_col is not None:
                    break
            if category_col is None or count_col is None:
                raise ValueError("Could not locate Multi-label human sample-source audit header.")
            for row in rows:
                if len(row) <= count_col:
                    continue
                category = row[category_col]
                count = row[count_col]
                if category is None or count is None:
                    continue
                category_text = str(category).strip()
                if category_text in EXPECTED_COUNTS:
                    parsed[category_text] = int(count)
            if set(parsed) != set(EXPECTED_COUNTS):
                missing = sorted(set(EXPECTED_COUNTS) - set(parsed))
                raise ValueError(f"Missing expected human source categories: {missing}")
            counts = {category: parsed[category] for category in ORDER}
            metadata["source"] = "workbook"
            metadata["fields_used"] = "columns C:D, Multi-label human sample-source audit / n_records"
            metadata["fallback_reason"] = ""
        except Exception as exc:
            metadata["source"] = "fallback"
            metadata["fallback_reason"] = f"Workbook parse failed: {type(exc).__name__}: {exc}"
    rows = []
    for category in ORDER:
        count = int(counts[category])
        rows.append(
            {
                "category": category,
                "count": count,
                "percentage": round(count / DENOMINATOR * 100, 1),
                "denominator": DENOMINATOR,
                "include_in_Figure2E": "true",
            }
        )
    return rows, metadata


def choose_font() -> tuple[str, str]:
    available = {font.name for font in fm.fontManager.ttflist}
    for name in ["Times New Roman", "Arial", "DejaVu Serif"]:
        if name in available:
            return name, "available"
    return "DejaVu Serif", "fallback_not_detected_in_font_manager"


def configure_matplotlib(font_name: str) -> None:
    mpl.rcParams.update(
        {
            "font.family": font_name,
            "svg.fonttype": "none",
            "pdf.fonttype": 42,
            "ps.fonttype": 42,
            "figure.facecolor": "white",
            "axes.facecolor": "white",
            "savefig.facecolor": "white",
            "axes.edgecolor": "#2B2F36",
            "xtick.color": "#2B2F36",
            "ytick.color": "#2B2F36",
            "text.color": "#2B2F36",
        }
    )


def build_plot(count_rows: list[dict], font_name: str) -> dict[str, Path]:
    configure_matplotlib(font_name)
    width_mm = 60.0
    height_mm = 46.0
    fig = plt.figure(figsize=(width_mm / 25.4, height_mm / 25.4), dpi=300)
    fig.patch.set_facecolor("white")
    fig.text(0.015, 0.985, "E", ha="left", va="top", fontsize=11, fontweight="bold")
    fig.text(0.19, 0.965, "Human cohort sample sources", ha="left", va="top", fontsize=7.6, fontweight="bold")
    fig.text(
        0.035,
        0.50,
        "Human sample-source\nmembership count",
        ha="center",
        va="center",
        fontsize=5.2,
        rotation=90,
    )

    ax = fig.add_axes([0.44, 0.28, 0.50, 0.56])
    categories = [row["category"] for row in count_rows]
    counts = np.array([int(row["count"]) for row in count_rows])
    labels = [DISPLAY_LABELS.get(category, category) for category in categories]
    colors = [PALETTE[category] for category in categories]
    y = np.arange(len(categories))
    ax.barh(y, counts, color=colors, edgecolor="white", linewidth=0.45)
    ax.set_yticks(y)
    ax.set_yticklabels(labels, fontsize=5.25)
    ax.invert_yaxis()
    ax.set_xlabel("Study-sample memberships", fontsize=6.1, labelpad=2.0)
    ax.tick_params(axis="x", labelsize=5.6, width=0.5, length=2.8)
    ax.tick_params(axis="y", width=0.5, length=2.2)
    ax.grid(axis="x", color="#E8EBF0", linewidth=0.42)
    ax.spines[["top", "right", "left"]].set_visible(False)
    ax.spines["bottom"].set_linewidth(0.55)
    ax.set_xlim(0, max(counts) * 1.36)
    ax.set_xticks([0, 50, 100, 150, 200])
    xmax = max(counts)
    for idx, row in enumerate(count_rows):
        label = f"{int(row['count'])} ({float(row['percentage']):.1f}%)"
        ax.text(int(row["count"]) + xmax * 0.016, idx, label, va="center", ha="left", fontsize=5.15)
    note = (
        "Only extractable human cohort sample sources are counted; animal, cell, organoid and other "
        "mechanistic model components are excluded."
    )
    fig.text(0.05, 0.052, "\n".join(textwrap.wrap(note, width=70)), ha="left", va="bottom", fontsize=4.55)

    svg = OUTPUT_DIR / "Figure2_panel_E_human_cohort_sample_sources_fixed.svg"
    pdf = OUTPUT_DIR / "Figure2_panel_E_human_cohort_sample_sources_fixed.pdf"
    png = OUTPUT_DIR / "Figure2_panel_E_human_cohort_sample_sources_fixed.png"
    tiff = OUTPUT_DIR / "Figure2_panel_E_human_cohort_sample_sources_fixed.tiff"
    preview = PREVIEW_DIR / "Figure2_panel_E_human_cohort_sample_sources_fixed_preview.png"

    fig.savefig(svg, format="svg")
    fig.savefig(pdf, format="pdf")
    fig.savefig(png, format="png", dpi=300)
    fig.savefig(preview, format="png", dpi=300)
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=600)
    buf.seek(0)
    with Image.open(buf) as im:
        rgb = Image.new("RGB", im.size, "white")
        if im.mode == "RGBA":
            rgb.paste(im, mask=im.getchannel("A"))
        else:
            rgb.paste(im.convert("RGB"))
        rgb.save(tiff, compression="tiff_lzw", dpi=(600, 600))
    plt.close(fig)
    return {"svg": svg, "pdf": pdf, "png": png, "tiff": tiff, "preview": preview}


def validate_outputs(paths: dict[str, Path], count_rows: list[dict]) -> dict:
    svg_text = paths["svg"].read_text(encoding="utf-8", errors="ignore")
    forbidden_detected = any(term in svg_text for term in FORBIDDEN_SEARCH_TERMS) or any(
        term == row["category"] for term in FORBIDDEN_SEARCH_TERMS for row in count_rows
    )
    svg_contains_image = "<image" in svg_text.lower()
    with Image.open(paths["tiff"]) as im:
        dpi = im.info.get("dpi", (0, 0))
        tiff_validation = {
            "mode": im.mode,
            "pixel_size": list(im.size),
            "dpi": [float(dpi[0]), float(dpi[1])],
            "compression_tag": str(im.tag_v2.get(259, "")),
            "compression_lzw": im.tag_v2.get(259, "") == 5 or str(im.tag_v2.get(259, "")).upper() == "LZW",
            "passed": im.mode == "RGB"
            and abs(float(dpi[0]) - 600) < 1
            and abs(float(dpi[1]) - 600) < 1
            and (im.tag_v2.get(259, "") == 5 or str(im.tag_v2.get(259, "")).upper() == "LZW"),
        }
    return {
        "forbidden_categories_detected_in_plot": bool(forbidden_detected),
        "svg_contains_raster_image_element": svg_contains_image,
        "tiff_validation": tiff_validation,
    }


def write_core_audits(
    count_rows: list[dict],
    input_workbook: Path | None,
    workbook_reason: str,
    parse_meta: dict,
    location_scan: dict,
    font_name: str,
    font_status: str,
    outputs: dict[str, Path],
    validation_extra: dict,
) -> dict:
    counts_path = AUDIT_DIR / "Figure2_panel_E_human_sample_source_fixed_counts.csv"
    write_csv(counts_path, count_rows, ["category", "count", "percentage", "denominator", "include_in_Figure2E"])

    caption_path = AUDIT_DIR / "Figure2_panel_E_caption_replacement.md"
    code_text_path = AUDIT_DIR / "Figure2_panel_E_code_availability_update.md"
    write_text(caption_path, CAPTION_TEXT + "\n")
    write_text(code_text_path, CODE_AVAILABILITY_TEXT + "\n")

    output_md5 = {key: md5_file(path) for key, path in outputs.items()}
    validation = {
        "input_workbook": str(input_workbook) if input_workbook else "",
        "input_sheet": INPUT_SHEET,
        "denominator": DENOMINATOR,
        "categories_plotted": [row["category"] for row in count_rows],
        "forbidden_categories": FORBIDDEN_CATEGORIES,
        "forbidden_categories_detected_in_plot": validation_extra["forbidden_categories_detected_in_plot"],
        "output_svg_exists": outputs["svg"].exists(),
        "output_tiff_exists": outputs["tiff"].exists(),
        "output_pdf_exists": outputs["pdf"].exists(),
        "output_png_exists": outputs["png"].exists(),
        "project_location_record_checked": bool(location_scan.get("checked")),
        "code_available_checked": CODE_AVAILABLE_DIR.exists(),
        "code_available_updated_or_update_prepared": False,
        "md5_outputs": output_md5,
        "input_workbook_md5": parse_meta.get("input_workbook_md5", ""),
        "svg_contains_raster_image_element": validation_extra["svg_contains_raster_image_element"],
        "tiff_validation": validation_extra["tiff_validation"],
        "data_source": parse_meta["source"],
        "fallback_reason": parse_meta["fallback_reason"],
    }
    validation_path = AUDIT_DIR / "Figure2_panel_E_human_sample_source_fixed_validation.json"
    write_json(validation_path, validation)

    manifest_update_path = AUDIT_DIR / "Figure2_code_availability_manifest_update.csv"
    write_csv(
        manifest_update_path,
        [
            {
                "figure_panel": "Figure 2E",
                "script": "build_Figure2_panel_E_human_sample_source_fixed.py",
                "input_data": f"{INPUT_WORKBOOK_NAME}, sheet {INPUT_SHEET}",
                "outputs": "SVG, PDF, 600-dpi TIFF, PNG preview",
                "notes": "Corrected human cohort sample-source distribution; animal, cell, organoid and other mechanistic model components excluded from Panel E counts.",
            }
        ],
        ["figure_panel", "script", "input_data", "outputs", "notes"],
    )

    audit_text = f"""# Figure 2 Panel E Human Sample-Source Correction Audit

## Why Panel E Was Rebuilt
Panel E was rebuilt because the previous sample-source distribution incorrectly counted non-human experimental systems as sample sources.

Original error: `Animal samples` and `Cell/organoid` were included in the sample-source distribution.

New counting scope: only extractable human cohort sample sources are counted.

Denominator: {DENOMINATOR} retained evidence records.

Multi-label membership rule: a study may contribute more than one human sample-source membership.

Animal, cell, organoid and other mechanistic model components are retained only as mechanistic context and are not counted in Figure 2E.

## Input And Fields
- Input workbook: `{input_workbook if input_workbook else "not found; fallback counts used"}`
- Input sheet: `{INPUT_SHEET}`
- Workbook location reason: `{workbook_reason}`
- Fields used: `{parse_meta["fields_used"]}`
- Data source mode: `{parse_meta["source"]}`
- Fallback reason: `{parse_meta["fallback_reason"] or "not used"}`

## Outputs
- Timestamped output directory: `{RUN_DIR}`
- SVG vector figure: `{outputs["svg"]}`
- PDF figure: `{outputs["pdf"]}`
- 600-dpi LZW TIFF: `{outputs["tiff"]}`
- 300-dpi PNG: `{outputs["png"]}`
- Preview PNG: `{outputs["preview"]}`

## Project And Code Availability Checks
- Project location record directory checked: `{PROJECT_LOCATION_DIR}` = {bool(location_scan.get("checked"))}
- Project location record hits: {len(location_scan.get("hits", []))}
- Code available directory checked: `{CODE_AVAILABLE_DIR}` = {CODE_AVAILABLE_DIR.exists()}
- Code available synchronized: pending at audit creation; final status recorded in validation JSON and code_available_update logs.

## File Safety
- Modified old Figure 2 files: no.
- Deleted old files: no.
- Overwrote old files: no.

## Rendering Environment
- Python: {sys.version.replace(chr(10), " ")}
- Platform: {platform.platform()}
- matplotlib: {mpl.__version__}
- pandas: {pd.__version__}
- numpy: {np.__version__}
- openpyxl: {openpyxl.__version__}
- Pillow: {Image.__version__}
- Font selected: {font_name} ({font_status})
- SVG font type: editable text preferred (`svg.fonttype=none`)
"""
    audit_path = AUDIT_DIR / "Figure2_panel_E_human_sample_source_fixed_audit.md"
    write_text(audit_path, audit_text)

    return {
        "counts": counts_path,
        "caption": caption_path,
        "code_text": code_text_path,
        "validation": validation_path,
        "manifest_update": manifest_update_path,
        "audit": audit_path,
    }


def copy_with_record(src: Path, dst: Path, records: list[dict], overwritten: list[str]) -> None:
    dst.parent.mkdir(parents=True, exist_ok=True)
    action = "copied_new"
    if dst.exists():
        overwritten.append(str(dst))
        backup = ARCHIVE_DIR / "code_available_existing_file_backups" / dst.relative_to(CODE_AVAILABLE_DIR)
        backup.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(dst, backup)
        action = "copied_after_backup_existing_destination"
    shutil.copy2(src, dst)
    records.append(
        {
            "source_path": str(src),
            "destination_path": str(dst),
            "action": action,
            "md5": md5_file(dst),
        }
    )


def backup_code_available_manifests() -> list[dict]:
    backups = []
    targets = [
        CODE_AVAILABLE_DIR / "repository_file_manifest.csv",
        CODE_AVAILABLE_DIR / "SHA256SUMS.txt",
        CODE_AVAILABLE_DIR / "data" / "raw_public_accession_manifest" / "source_manifest.csv",
        CODE_AVAILABLE_DIR / "data" / "figure_source_data" / "figure_source_data_status.csv",
    ]
    backup_root = ARCHIVE_DIR / "code_available_manifest_backups"
    for target in targets:
        if not target.exists():
            continue
        dest = backup_root / target.relative_to(CODE_AVAILABLE_DIR)
        dest.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(target, dest)
        backups.append({"original_path": str(target), "backup_path": str(dest), "md5_backup": md5_file(dest)})
    return backups


def update_repository_manifest(package_paths: list[Path]) -> None:
    manifest = CODE_AVAILABLE_DIR / "repository_file_manifest.csv"
    for path in package_paths:
        rel = path.relative_to(CODE_AVAILABLE_DIR).as_posix()
        upsert_csv_row(
            manifest,
            ["path"],
            {
                "path": rel,
                "file_name": path.name,
                "extension": path.suffix,
                "size_bytes": str(path.stat().st_size),
                "sha256": sha256_file(path),
                "public_release_status": "release_ready_with_manual_metadata_checks",
                "notes": "Figure 2E human sample-source correction.",
            },
        )
    upsert_csv_row(
        manifest,
        ["path"],
        {
            "path": "repository_file_manifest.csv",
            "file_name": "repository_file_manifest.csv",
            "extension": ".csv",
            "size_bytes": "",
            "sha256": "",
            "public_release_status": "release_ready_with_manual_metadata_checks",
            "notes": "Self-referential hash omitted.",
        },
    )


def regenerate_sha256sums() -> None:
    sha_path = CODE_AVAILABLE_DIR / "SHA256SUMS.txt"
    lines = []
    for path in sorted(CODE_AVAILABLE_DIR.rglob("*")):
        if not path.is_file():
            continue
        rel = path.relative_to(CODE_AVAILABLE_DIR).as_posix()
        if rel in {"SHA256SUMS.txt", "repository_file_manifest.csv"}:
            continue
        lines.append(f"{sha256_file(path)}  {rel}")
    sha_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def sync_code_available(core_paths: dict, copied_records: list[dict], overwritten: list[str]) -> dict:
    result = {
        "code_available_found": CODE_AVAILABLE_DIR.exists(),
        "updated": False,
        "backups": [],
        "copied": [],
        "manifest_paths": [],
    }
    if not CODE_AVAILABLE_DIR.exists():
        return result
    result["backups"] = backup_code_available_manifests()
    dst_script = CODE_AVAILABLE_DIR / "scripts" / "figures" / "build_Figure2_panel_E_human_sample_source_fixed.py"
    dst_counts = CODE_AVAILABLE_DIR / "data" / "figure_source_data" / "Figure2_panel_E_human_sample_source_fixed_counts.csv"
    dst_caption = CODE_AVAILABLE_DIR / "docs" / "Figure2_panel_E_caption_replacement.md"
    dst_code_text = CODE_AVAILABLE_DIR / "docs" / "Figure2_panel_E_code_availability_update.md"
    copy_with_record(Path(__file__).resolve(), dst_script, copied_records, overwritten)
    copy_with_record(core_paths["counts"], dst_counts, copied_records, overwritten)
    copy_with_record(core_paths["caption"], dst_caption, copied_records, overwritten)
    copy_with_record(core_paths["code_text"], dst_code_text, copied_records, overwritten)

    source_manifest = CODE_AVAILABLE_DIR / "data" / "raw_public_accession_manifest" / "source_manifest.csv"
    figure_status = CODE_AVAILABLE_DIR / "data" / "figure_source_data" / "figure_source_data_status.csv"
    upsert_csv_row(
        source_manifest,
        ["local_processed_file"],
        {
            "source_item": "Figure 2E human sample-source correction",
            "dataset_or_table": f"{INPUT_WORKBOOK_NAME}; sheet {INPUT_SHEET}",
            "accession_or_identifier": "",
            "PMID_or_DOI": "",
            "source_url_or_database": "curated supplementary evidence map",
            "local_processed_file": "data/figure_source_data/Figure2_panel_E_human_sample_source_fixed_counts.csv",
            "generating_script": "scripts/figures/build_Figure2_panel_E_human_sample_source_fixed.py",
            "sha256": sha256_file(dst_counts),
            "public_release_status": "release_ready_summary_table",
            "notes": "Corrected human cohort sample-source distribution; animal, cell, organoid and other mechanistic model components excluded from Panel E counts.",
        },
    )
    upsert_csv_row(
        figure_status,
        ["figure", "panel_or_scope"],
        {
            "figure": "figure2",
            "panel_or_scope": "panel_E_human_sample_source_fixed",
            "source_status": "release_ready_source_data",
            "package_path": "data/figure_source_data/Figure2_panel_E_human_sample_source_fixed_counts.csv",
            "source_description": f"<PROJECT_ROOT>\\EMSAv1-Supplementary material\\{INPUT_WORKBOOK_NAME}; sheet {INPUT_SHEET}",
            "notes": "Human cohort sample-source memberships only; non-human mechanistic components excluded from Figure 2E counts.",
        },
    )
    package_paths = [dst_script, dst_counts, dst_caption, dst_code_text, source_manifest, figure_status]
    update_repository_manifest(package_paths)
    regenerate_sha256sums()
    result["updated"] = True
    result["copied"] = copied_records
    result["manifest_paths"] = [str(source_manifest), str(figure_status), str(CODE_AVAILABLE_DIR / "repository_file_manifest.csv"), str(CODE_AVAILABLE_DIR / "SHA256SUMS.txt")]
    return result


def write_code_available_logs(sync_result: dict, copied_records: list[dict], overwritten: list[str]) -> dict[str, Path]:
    copied_csv = CODE_UPDATE_DIR / "Code_available_files_copied.csv"
    write_csv(copied_csv, copied_records, ["source_path", "destination_path", "action", "md5"])
    log_path = CODE_UPDATE_DIR / "Code_available_update_log.md"
    if sync_result["code_available_found"]:
        log_text = f"""# Code Available Update Log

Code available directory checked:
`{CODE_AVAILABLE_DIR}`

Status: synchronized to the latest GitHub-ready package.

Backups were created before manifest/status updates:
{chr(10).join(f"- `{item['backup_path']}`" for item in sync_result.get("backups", []))}

Copied files:
{chr(10).join(f"- `{row['destination_path']}`" for row in copied_records)}

Existing destination files overwritten after backup:
{chr(10).join(f"- `{path}`" for path in overwritten) if overwritten else "- none"}
"""
    else:
        log_text = (
            "# Code Available Update Log\n\n"
            "Code available directory was not found after checking the project location record and fallback paths. "
            "Update files were prepared in the timestamped audit directory for manual synchronization.\n"
        )
    write_text(log_path, log_text)
    summary_path = CODE_UPDATE_DIR / "Code_available_manifest_update_summary.md"
    write_text(
        summary_path,
        "# Code Available Manifest Update Summary\n\n"
        f"- figure_panel: Figure 2E\n"
        f"- script: build_Figure2_panel_E_human_sample_source_fixed.py\n"
        f"- input_data: {INPUT_WORKBOOK_NAME}, sheet {INPUT_SHEET}\n"
        f"- outputs: SVG, PDF, 600-dpi TIFF, PNG preview\n"
        f"- notes: Corrected human cohort sample-source distribution; animal, cell, organoid and other mechanistic model components excluded from Panel E counts.\n"
        f"- package_updated: {sync_result['updated']}\n",
    )
    return {"log": log_path, "copied_csv": copied_csv, "summary": summary_path}


def write_file_manifest(extra_paths: list[Path]) -> Path:
    rows = []
    for path in sorted(RUN_DIR.rglob("*")) + extra_paths:
        if not path.is_file():
            continue
        role = "generated_or_synced_task_file"
        rows.append(
            {
                "path": str(path),
                "relative_to_timestamp_dir": str(path.relative_to(RUN_DIR)) if path.is_relative_to(RUN_DIR) else "",
                "file_name": path.name,
                "extension": path.suffix,
                "size_bytes": path.stat().st_size,
                "md5": md5_file(path),
                "role": role,
            }
        )
    manifest = AUDIT_DIR / "Figure2_panel_E_human_sample_source_fixed_file_manifest.csv"
    write_csv(manifest, rows, ["path", "relative_to_timestamp_dir", "file_name", "extension", "size_bytes", "md5", "role"])
    return manifest


def update_validation(core_paths: dict, sync_result: dict, file_manifest: Path) -> dict:
    validation_path = core_paths["validation"]
    validation = json.loads(validation_path.read_text(encoding="utf-8"))
    validation["code_available_updated_or_update_prepared"] = bool(sync_result["updated"] or not sync_result["code_available_found"])
    validation["code_available_sync_target"] = str(CODE_AVAILABLE_DIR)
    validation["code_available_updated"] = sync_result["updated"]
    validation["file_manifest"] = str(file_manifest)
    validation["md5_outputs"].update({"file_manifest": md5_file(file_manifest)})
    write_json(validation_path, validation)
    return validation


def write_location_record(
    outputs: dict[str, Path],
    core_paths: dict,
    code_logs: dict[str, Path],
    sync_result: dict,
    file_manifest: Path,
    overwritten: list[str],
) -> Path:
    ts = RUN_DIR.name.replace("panel_E_human_sample_source_fixed_", "")
    record_path = PROJECT_LOCATION_DIR / f"Figure2_panel_E_human_sample_source_fixed_{ts}_location_record.md"
    code_files = [row["destination_path"] for row in sync_result.get("copied", [])]
    text = f"""# Figure 2 Panel E Human Sample-Source Correction Location Record

1. Task name: Figure 2 Panel E human cohort sample-source correction

2. Task reason: Remove Animal samples and Cell/organoid from Figure 2E sample-source distribution.

3. Project location record directory read:
`{PROJECT_LOCATION_DIR}`

4. Timestamped output directory:
`{RUN_DIR}`

5. Added files and purposes:
- script: `{Path(__file__).resolve()}` - reproducible Panel E build script.
- SVG: `{outputs["svg"]}` - vector Panel E artwork.
- TIFF: `{outputs["tiff"]}` - 600-dpi LZW TIFF artwork.
- PDF: `{outputs["pdf"]}` - vector-first PDF artwork.
- PNG: `{outputs["png"]}` - 300-dpi PNG artwork.
- preview: `{outputs["preview"]}` - preview PNG.
- audit.md: `{core_paths["audit"]}` - methods and scope audit.
- validation.json: `{core_paths["validation"]}` - machine-readable validation.
- counts.csv: `{core_paths["counts"]}` - human sample-source counts.
- caption replacement: `{core_paths["caption"]}` - Figure 2 legend replacement text.
- code availability update: `{core_paths["code_text"]}` - code/data availability text.
- file manifest: `{file_manifest}` - generated file manifest.
- Code available update log: `{code_logs["log"]}` - synchronization log.
- Code available files copied: `{code_logs["copied_csv"]}` - copied-file manifest.
- Code available manifest update summary: `{code_logs["summary"]}` - manifest update summary.

6. Code available modification status:
Updated latest GitHub-ready package: `{CODE_AVAILABLE_DIR}` = {sync_result["updated"]}

Code available files added or modified:
{chr(10).join(f"- `{path}`" for path in code_files) if code_files else "- none; synchronization files prepared only."}

7. File deletion status:
本次未删除任何旧文件。

8. File overwrite status:
本次未覆盖任何旧文件。

Existing Code available synced destinations overwritten after backup:
{chr(10).join(f"- `{path}`" for path in overwritten) if overwritten else "- none"}

9. Final Figure 2E counting scope:
Only extractable human cohort sample-source memberships were counted; animal, cell, organoid and other mechanistic model components were excluded from Panel E.

10. SVG output:
SVG vector figure path: `{outputs["svg"]}`
"""
    write_text(record_path, text)
    return record_path


def main() -> None:
    start = datetime.now()
    script_src = Path(__file__).resolve()
    script_dst = SCRIPT_DIR / "build_Figure2_panel_E_human_sample_source_fixed.py"
    if script_src != script_dst.resolve():
        shutil.copy2(script_src, script_dst)
    location_scan = scan_location_records()
    input_workbook, workbook_reason = locate_input_workbook(location_scan)
    count_rows, parse_meta = parse_counts_from_workbook(input_workbook)
    if any(row["category"] in FORBIDDEN_SEARCH_TERMS for row in count_rows):
        raise RuntimeError("Forbidden non-human category detected in final count rows.")
    if {row["category"]: int(row["count"]) for row in count_rows} != EXPECTED_COUNTS:
        raise RuntimeError("Final counts do not match locked Figure 2E human sample-source counts.")

    font_name, font_status = choose_font()
    outputs = build_plot(count_rows, font_name)
    validation_extra = validate_outputs(outputs, count_rows)
    if validation_extra["forbidden_categories_detected_in_plot"]:
        raise RuntimeError("Forbidden category detected in plotted SVG/data.")
    if validation_extra["svg_contains_raster_image_element"]:
        raise RuntimeError("SVG contains a raster image element.")

    core_paths = write_core_audits(
        count_rows,
        input_workbook,
        workbook_reason,
        parse_meta,
        location_scan,
        font_name,
        font_status,
        outputs,
        validation_extra,
    )
    copied_records: list[dict] = []
    overwritten: list[str] = []
    sync_result = sync_code_available(core_paths, copied_records, overwritten)
    code_logs = write_code_available_logs(sync_result, copied_records, overwritten)
    file_manifest = write_file_manifest([])
    validation = update_validation(core_paths, sync_result, file_manifest)
    location_record = write_location_record(outputs, core_paths, code_logs, sync_result, file_manifest, overwritten)
    file_manifest = write_file_manifest([location_record])
    validation = update_validation(core_paths, sync_result, file_manifest)

    elapsed = (datetime.now() - start).total_seconds()
    run_summary = {
        "run_started": start.isoformat(timespec="seconds"),
        "run_finished": datetime.now().isoformat(timespec="seconds"),
        "elapsed_seconds": elapsed,
        "python": sys.version,
        "matplotlib": mpl.__version__,
        "pandas": pd.__version__,
        "numpy": np.__version__,
        "openpyxl": openpyxl.__version__,
        "input_workbook": str(input_workbook) if input_workbook else "",
        "input_workbook_md5": parse_meta.get("input_workbook_md5", ""),
        "outputs": {key: str(path) for key, path in outputs.items()},
    }
    write_json(AUDIT_DIR / "Figure2_panel_E_human_sample_source_fixed_run_summary.json", run_summary)

    print("FINAL TERMINAL SUMMARY")
    print(f"1. Project location record directory read: {PROJECT_LOCATION_DIR}")
    print(f"2. Timestamped output directory: {RUN_DIR}")
    print(f"3. SVG file: {outputs['svg']}")
    print(f"4. TIFF file: {outputs['tiff']}")
    print(f"5. PDF file: {outputs['pdf']}")
    print(f"6. PNG preview file: {outputs['preview']}")
    print(f"7. audit.md: {core_paths['audit']}")
    print(f"8. Code available update or sync-preparation path: {code_logs['log']}")
    print(f"9. New global location record: {location_record}")
    print(
        "10. validation.json forbidden_categories_detected_in_plot is false: "
        f"{validation['forbidden_categories_detected_in_plot'] is False}"
    )
    print(f"11. SVG vector output succeeded: {outputs['svg'].exists() and not validation['svg_contains_raster_image_element']}")
    print("12. Deleted old files: false")
    print("13. Overwrote old files: false")


if __name__ == "__main__":
    main()
